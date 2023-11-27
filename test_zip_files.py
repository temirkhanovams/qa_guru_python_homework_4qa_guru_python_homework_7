import os.path
import zipfile
import openpyxl
import xlrd
from utils import *
from pypdf import PdfReader


def test_names_of_files():
    with zipfile.ZipFile(path_tmp_zip) as zf:
        file_names_in_resources = os.listdir(path_resources)
        file_names_zip_in_tmp = zf.namelist()
    for file in file_names_in_resources:
        assert file in file_names_zip_in_tmp


def test_sizes_of_files():
    with zipfile.ZipFile(path_tmp_zip) as zf:
        for file in zf.namelist():
            print('\n', file, '\n')
            original_file = os.path.join(path_resources, file)
            assert zf.getinfo(file).file_size == os.path.getsize(original_file)
            with open(original_file, 'rb') as f:
                assert zf.read(file) == f.read()


def test_xls():
    file_name = "file_example_XLS_10.xls"
    book_in_resources = xlrd.open_workbook(os.path.join(path_resources, file_name))
    with zipfile.ZipFile(path_tmp_zip) as zf:
        book_in_zip = xlrd.open_workbook(file_contents=zf.read(file_name))
        assert book_in_zip.nsheets == book_in_resources.nsheets
        assert book_in_zip.sheet_names() == book_in_resources.sheet_names()
        assert book_in_zip.sheet_by_index(0).ncols == book_in_resources.sheet_by_index(0).ncols
        assert book_in_zip.sheet_by_index(0).nrows == book_in_resources.sheet_by_index(0).nrows
        assert book_in_zip.sheet_by_index(0).col_values(1) == book_in_resources.sheet_by_index(0).col_values(1)
        assert book_in_zip.sheet_by_index(0).row_values(0) == book_in_resources.sheet_by_index(0).row_values(0)
        assert book_in_zip.sheet_by_index(0).cell_value(1, 2) == book_in_resources.sheet_by_index(0).cell_value(1, 2)


def test_xlsx():
    file_name = "file_example_XLSX_50.xlsx"
    book_in_resources = openpyxl.load_workbook(os.path.join(path_resources, file_name))
    sheet_in_resources = book_in_resources.active
    with zipfile.ZipFile(path_tmp_zip) as zf:
        book_in_zip = openpyxl.load_workbook(zf.open(file_name))
        assert book_in_zip.sheetnames == book_in_resources.sheetnames


def test_pdf():
    file_name = "Python Testing with Pytest (Brian Okken).pdf"
    pdf_in_resources = PdfReader(os.path.join(path_resources, file_name))
    with zipfile.ZipFile(path_tmp_zip) as zip_file:
        pdf_in_zip = PdfReader(zip_file.open(file_name, 'r'))
        assert len(pdf_in_zip.pages) == len(pdf_in_resources.pages)
        assert pdf_in_zip.pages[10].extract_text() == pdf_in_resources.pages[10].extract_text()
        assert not pdf_in_zip.attachments